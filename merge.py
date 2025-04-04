# -*- coding: utf-8 -*-
import openpyxl
import copy
import glob
import os
import sys

# 修改目录work_dir
work_dir = os.path.dirname(os.path.abspath(__file__))

# 检查命令行参数
if len(sys.argv) > 1:
    # 如果提供了文件参数，则使用这些文件
    files_to_merge = []
    output_file = None
    
    # 处理命令行参数
    for arg in sys.argv[1:]:
        if arg.lower().endswith('.xlsx'):
            if not output_file:
                files_to_merge.append(os.path.join(work_dir, arg) if not os.path.isabs(arg) else arg)
            else:
                output_file = os.path.join(work_dir, arg) if not os.path.isabs(arg) else arg
else:
    # 否则使用目录中的所有xlsx文件
    file_name = '*.xlsx'
    files_to_merge = [f for f in glob.glob(os.path.join(work_dir, file_name)) if not f.endswith('merged.xlsx')]

if not files_to_merge:
    print("没有找到可以合并的Excel文件！")
    sys.exit(1)

# 创建一个新的工作表
new_wb = openpyxl.Workbook()
new_sheet = new_wb.create_sheet('Merged',0)

# 遍历所有excel文件的sheet,存为list
wb_list = []
sheet_list = []
for f in files_to_merge:
    try:
        wb1 = openpyxl.load_workbook(f, data_only=True)
        # 只获取第一个sheet
        sheet = wb1.active
        wb_list.append(wb1)
        sheet_list.append(sheet)
    except Exception as e:
        print(f"无法打开文件 {f}: {str(e)}")
        continue

# 合并所有sheet中的数据，带格式，复制到新的工作表中
row_begin = 0
for i, sheet in enumerate(sheet_list):
    print(f"正在处理文件: {files_to_merge[i]}")
    
    # 复制数据和格式
    for n_r, row in enumerate(sheet.rows):
        for n_c, source_cell in enumerate(row):
            target_cell = new_sheet.cell(row=row_begin+n_r+1, column=n_c+1)
            
            # 处理合并单元格的值
            if isinstance(source_cell, openpyxl.cell.cell.MergedCell):
                # 获取合并单元格的主单元格值
                for merged_range in sheet.merged_cells:
                    if source_cell.coordinate in merged_range:
                        main_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        target_cell.value = main_cell.value
                        break
            else:
                target_cell.value = source_cell.value

            # 复制样式（如果源单元格有样式）
            if hasattr(source_cell, 'has_style') and source_cell.has_style:
                
                # 然后复制其他样式，但保持换行属性
                target_cell._style = copy.copy(source_cell._style)
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = copy.copy(source_cell.number_format)
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)

    # 处理当前sheet的合并单元格
    for merged_range in sheet.merged_cells:
        new_start_row = merged_range.min_row + row_begin
        new_end_row = merged_range.max_row + row_begin
        new_range = f"{openpyxl.utils.get_column_letter(merged_range.min_col)}{new_start_row}:{openpyxl.utils.get_column_letter(merged_range.max_col)}{new_end_row}"
        try:
            new_sheet.merge_cells(new_range)
        except ValueError:
            pass  # 忽略已经合并的单元格

    # 更新下一个文件的起始行
    row_begin += sheet.max_row

# 复制条件格式
for cf in sheet.conditional_formatting:
    new_cf = copy.copy(cf)
    # 调整条件格式的范围
    old_ranges = cf.cells.ranges
    new_ranges = []
    for old_range in old_ranges:
        boundaries = openpyxl.utils.cell.range_boundaries(str(old_range))
        start_row = boundaries[0] + row_begin
        end_row = boundaries[2] + row_begin
        new_range = f"{openpyxl.utils.get_column_letter(boundaries[1])}{start_row}:{openpyxl.utils.get_column_letter(boundaries[3])}{end_row}"
        new_ranges.append(new_range)
    new_cf.cells.ranges = new_ranges
    new_sheet.conditional_formatting.append(new_cf)

# Get height of row 1 (default to 15 if not set)
row1_height = new_sheet.row_dimensions[1].height or 15

# Set rows 3-6 to double height
for row in range(3, 7):
    new_sheet.row_dimensions[row].height = row1_height * 2

# Save the new Excel file
merged_file = os.path.join(work_dir,'merged.xlsx')
new_wb.save(merged_file)
print("save excel to: " + merged_file)

# 在Windows系统下自动打开合并后的Excel文件
# if os.name == 'nt':
#     os.startfile(merged_file)
#     print("Opening merged Excel file...")
#     print("按回车键退出程序...")
#     input()