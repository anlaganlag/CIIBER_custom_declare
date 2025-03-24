from inspect import BufferFlags
import pandas as pd
import os
import argparse
import sys
import subprocess
from openpyxl import load_workbook

# =============================================================================
# Configuration Section
# =============================================================================
# Default configuration values used if config.py is not available
# These lists and dictionaries define how data should be processed and mapped

# Columns to preserve from the input file (green headers in the specification)
PRESERVED_COLUMNS = ['Column1', 'Column2', 'Column3']

# Column containing material codes used for lookup/matching between files
MATERIAL_CODE_COLUMN = 'MaterialCode'

# Columns to be matched based on material code (yellow headers in the specification)
MATCHED_COLUMNS = ['MatchedColumn1', 'MatchedColumn2']

# Columns with fixed/static values to be added to the output
FIXED_COLUMNS = {'FixedColumn1': 'Fixed Value 1', 'FixedColumn2': 'Fixed Value 2'}


fill_dict = {
    "境内发货人": "发票卖方",
    "境外收货人": "发票买方",
    "生产销售单位": "发票卖方",
    "合同协议号": "出口发票号",
    # "监管方式": "一般贸易",
    # "征免性质": "一般征税",
    # "贸易国(地区)": "印度",
    # "运抵国(地区)": "印度",
    "件数": "PL的件数",
    "净重(千克)": "PL的净重",
    "毛重(千克)": "PL的毛重",
    "运费（CNY)": "运价表的总运费",
    "保费（CNY)": "运价表的总保费"
}

# Column name mapping from English to Chinese
# Used to translate column headers between different languages
COLUMN_MAPPING = {
    'NO.': '项号',           # Item number
    'DESCRIPTION': '品名',   # Product name
    'Model NO.': '型号',     # Model number
    'Qty': '数量',           # Quantity
    'Unit': '单位',          # Unit
    'Amount': '总价',        # Total price
    'net weight': '净重',    # Net weight
    'Unit Price': '单价',    # Unit price
}

# Try to import custom configuration from config.py
# If the file exists, it will override the default values above
try:
    from config import PRESERVED_COLUMNS, MATERIAL_CODE_COLUMN, MATCHED_COLUMNS, FIXED_COLUMNS
    # Note: COLUMN_MAPPING is not imported from config.py and will always use the default
except ImportError:
    print("Warning: config.py file not found. Using default configuration.")

def convert_excel(input_file, reference_file, output_file):
    """
    Convert Excel file according to specified requirements.
    
    This function processes data from an input Excel file, matches it with data
    from a reference file, and produces a new Excel file with the transformed data.
    The transformation includes:
    - Copying specified columns from the input file
    - Matching data with the reference file based on material codes
    - Adding fixed value columns
    - Reordering columns according to a predefined order
    
    Args:
        input_file (str): Path to the first Excel file (source data)
        reference_file (str): Path to the reference Excel file (for material code matching)
        output_file (str): Path to save the output Excel file
        
    Returns:
        pandas.DataFrame: The processed DataFrame that was saved to the output file
        None: If an error occurred during conversion
    """
    # Check if input files exist
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found.")
        print(f"Current directory: {os.getcwd()}")
        print("Available files:")
        for file in os.listdir('.'):
            if file.endswith('.xlsx'):
                print(f"  - {file}")
        return None
        
    if not os.path.exists(reference_file):
        print(f"Error: Reference file '{reference_file}' not found.")
        print(f"Current directory: {os.getcwd()}")
        print("Available files:")
        for file in os.listdir('.'):
            if file.endswith('.xlsx'):
                print(f"  - {file}")
        return None
    
    # Read the input Excel file
    print(f"Reading input file: {input_file}")
    
    try:
        # Get the number of sheets in the Excel file
        excel_file = pd.ExcelFile(input_file)
        sheet_count = len(excel_file.sheet_names)
        
        # Choose the appropriate sheet based on sheet count
        # If there are 2 or more sheets, use the second sheet (index 1)
        # Otherwise, use the first sheet (index 0)
        sheet_to_read = 1 if sheet_count >= 2 else 0
        df_input = pd.read_excel(input_file, skiprows=9, sheet_name=sheet_to_read)
    except Exception as e:
        print(f"Error reading input file: {e}")
        return None
    
    # Data cleaning operations
    # =======================
    
    # Safely delete row 0 (if it exists) and reset index
    # This is often necessary when Excel files have header rows that aren't part of the data
    if len(df_input) > 0:  # Check if DataFrame is not empty
        df_input = df_input.drop(index=0).reset_index(drop=True)
    
    # Strip whitespace from column names only if DataFrame is not empty and has columns
    if not df_input.empty and len(df_input.columns) > 0:
        df_input.columns = df_input.columns.str.strip()
    
    # Strip whitespace from string data in all columns
    # Note: This loop only iterates through object (string) columns
    for column in df_input.select_dtypes(include=['object']).columns:
        df_input[column] = df_input[column]
    
    # Find the first empty NO. row and filter the dataframe
    # This assumes that data after the first empty NO. row should be ignored
    if 'NO.' in df_input.columns:
        # Convert NO. column to string and strip whitespace
        df_input['NO.'] = df_input['NO.'].astype(str).str.strip()
        
        # Find the first empty NO. row (containing 'nan', '', or ' ')
        empty_no_index = df_input[df_input['NO.'].isin(['nan', '', ' '])].index
        if len(empty_no_index) > 0:
            first_empty_index = empty_no_index[0]
            # Keep only rows before the first empty NO.
            df_input = df_input.iloc[:first_empty_index].copy()
    
    # Print columns found in the input file for debugging
    print(f"Input file columns: {df_input.columns.tolist()}")
    
    # Read the reference Excel file used for matching material codes
    print(f"Reading reference file: {reference_file}")
    df_reference = pd.read_excel(reference_file)
    
    # Create a new DataFrame for the output
    df_output = pd.DataFrame()
    
    # Define the desired column order for the output file
    # These are the required columns in the final output with Chinese headers
    column_order = [
        '项号',              # Item number
        '商品编号',          # Product code
        '品名',              # Product name
        '型号',              # Model number
        '申报要素',          # Declaration elements
        '数量',              # Quantity
        '单位',              # Unit
        '单价',              # Unit price
        '总价',              # Total price
        '币制',              # Currency
        '原产国（地区）',    # Country (region) of origin
        '最终目的国（地区）', # Final destination country (region)
        '境内货源地',        # Domestic source
        '征免',              # Tax exemption
        '净重'               # Net weight
    ]
    
    # Copy preserved columns (green headers) from input file with Chinese column names
    # This loop matches English column names from the input file to their Chinese equivalents
    for col in PRESERVED_COLUMNS:
        for eng_col, cn_col in COLUMN_MAPPING.items():
            if eng_col in df_input.columns and cn_col == col:
                df_output[col] = df_input[eng_col]
                break
        else:
            # This else clause is executed if the break is not reached (column not found)
            print(f"Warning: Column '{col}' not found in input file")
    
    # Match columns by material code (yellow headers)
    # First, find the English column name that corresponds to the material code column
    material_code_eng = next((k for k, v in COLUMN_MAPPING.items() if v == MATERIAL_CODE_COLUMN), MATERIAL_CODE_COLUMN)
    print(f"Looking for material code column: {material_code_eng} or {MATERIAL_CODE_COLUMN}")
    
    # Check if the material code columns exist in both files before attempting to match
    if material_code_eng in df_input.columns and MATERIAL_CODE_COLUMN in df_reference.columns:
        print("Found material code columns in both files")
        
        # Create a mapping dictionary for faster lookups
        # This avoids expensive DataFrame merges for each matched column
        reference_dict = {}
        for col in MATCHED_COLUMNS:
            if col.upper() == '商品编号':
                col = "HSCODE"
            if col in df_reference.columns:
                print(f"Creating mapping for column '{col}'")
                reference_dict[col] = df_reference.set_index(MATERIAL_CODE_COLUMN)[col].to_dict()
            else:
                print(f"Warning: Matched column '{col}' not found in reference file")
        
        # Add matched columns to output DataFrame using the dictionary mapping
        for col in MATCHED_COLUMNS:
            if col == '商品编号':
                col = "HSCODE"
            if col in reference_dict:
                print(f"Applying mapping for column '{col}'")
                if col == "HSCODE":
                    df_output["商品编号"] = df_input[material_code_eng].map(reference_dict[col])
                # Use the material code from input to look up values in the reference dictionary
                df_output[col] = df_input[material_code_eng].map(reference_dict[col])
    else:
        # Print detailed error information if material code columns are not found
        print(f"Warning: Material code column not found in one of the files")
        print(f"Input columns available: {df_input.columns.tolist()}")
        print(f"Reference columns available: {df_reference.columns.tolist()}")
    
    # Add fixed columns with static values
    for col, value in FIXED_COLUMNS.items():
        print(f"Adding fixed column '{col}' with value '{value}'")
        df_output[col] = value
    
    # Reorder columns according to the desired order
    print("Reordering columns according to specified order")
    print(f"Column order: {column_order}")
    df_output = df_output.reindex(columns=column_order)
    print(f"Final columns: {df_output.columns.tolist()}")
    
    # Save the output Excel file
    print(f"Saving output file: {output_file}")
    
    # Create a new Excel writer object
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write the main data
        df_output.to_excel(writer, index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Sheet1']
        
        # Set column widths
        for idx, col in enumerate(df_output.columns):
            worksheet.column_dimensions[chr(65 + idx)].width = 15
    
    print("Conversion completed successfully!")
    
    # 处理1.xlsx文件中的件数、毛重和净重信息
    
    try:
        print("Processing input.xlsx(PL) for TTL data...")
        df_1 = pd.read_excel('input.xlsx', sheet_name=0)
        
        # 初始化变量
        cnt = gw = nw = 0
        
        # 遍历A列查找'TTL:'
        for idx, value in enumerate(df_1.iloc[:, 0]):
            if isinstance(value, str) and value.strip() == 'TTL:':
                # 检查上一行是否为数字
                prev_value = df_1.iloc[idx-1, 0] if idx > 0 else None
                if isinstance(prev_value, (int, float)):
                    # 获取第6、8、9列的值（索引为5、7、8）
                    cnt = float(df_1.iloc[idx, 5])  if pd.notna(df_1.iloc[idx, 5]) else 0
                    gw = float(df_1.iloc[idx, 7])  if pd.notna(df_1.iloc[idx, 7]) else 0
                    nw = float(df_1.iloc[idx, 8])  if pd.notna(df_1.iloc[idx, 8]) else 0
                    print(f"Found TTL data: cnt={cnt}, gw={gw}, nw={nw}")
                    break
        fill_dict["件数"] = str(cnt)
        fill_dict["毛重(千克)"] = str(gw)
        fill_dict["净重(千克)"] = str(nw)


    except Exception as e:
        print(f"Error processing input(PL).xlsx for weight and quantity information: {e}")

    # 处理input.xlsx(发票)文件中的境内发货人,境外收货人,生产销售单位,合同协议号
    try:
        seller = buyer = no = ""

        from openpyxl import load_workbook

        # 直接读取 Excel 文件
        wb = load_workbook("input.xlsx")
        ws = wb.worksheets[1]  # sheet_name=1 对应第二个工作表

        # 直接获取 A1 单元格的值
        seller = ws["A1"].value
        fill_dict['境内发货人'] = seller
        fill_dict['生产销售单位'] = seller

        df_1 = pd.read_excel('input.xlsx', sheet_name=1)

        
        # 读取前四行所有单元格查找buyer和CI No.
        for i in range(4):
            for j in range(len(df_1.columns)):
                # 读取当前单元格的值
                cell_value = df_1.iloc[i, j] if not pd.isna(df_1.iloc[i, j]) else ""
                if not cell_value:  # 如果是空字符串则跳过
                    continue
                cell_value = str(cell_value)
                
                # 检查是否包含"Buyer:"
                if "Buyer:" in cell_value:
                    # 直接从当前单元格提取冒号后面的内容
                    buyer = cell_value.split(":", 1)[1].strip()
                
                # 检查是否包含"CI No.:"
                if "CI No.:" in cell_value:
                    # 如果在同一行找到CI No.信息，获取下一列的值
                    next_col = df_1.iloc[i, j+1] if j+1 < len(df_1.columns) and not pd.isna(df_1.iloc[i, j+1]) else ""
                    no = next_col if next_col else ""
        fill_dict['境外收货人'] = buyer
        fill_dict["合同协议号"] = no
        print(f"Extracted info - Seller: {seller}, Buyer: {buyer}, CI No.: {no}")

    except Exception as e:
        print(f"处理发票信息时出错: {e}")


    # 计算总货值和总净重
    t_amount = round(df_output['总价'].sum(), 2) if '总价' in df_output.columns else 0
    t_weight = round(df_output['净重'].sum(), 2) if '净重' in df_output.columns else 0
    exchange_rate = 0.139275766016713
    shipping_rate = 1.795242141
    t_insurance =round( (t_amount * 1.05*1.1*0.0005 ) /exchange_rate,2)
    t_shipping = round( t_weight * shipping_rate,2)
    fill_dict["运费（CNY)"] = str(t_shipping)
    fill_dict["保费（CNY)"] = str(t_insurance)

    # 处理1.xlsx文件的件数、毛重和净重信息
    try:
        print("Processing 1.xlsx for weight and quantity information...")
        if os.path.exists('1.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook('1.xlsx')
            ws = wb.active

        
        # 遍历前10行查找并修改特定单元格
            for row in range(1, 11):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        if "件数" in cell.value:
                            cell.value = f"件数 \n{cnt}"
                        elif "毛重(千克)" in cell.value:
                            cell.value = f"毛重(千克)\n{gw}"
                        elif "净重(千克)" in cell.value:
                            cell.value = f"净重(千克)\n{nw}"
                        elif "监管方式" in cell.value:
                            cell.value = f"监管方式\n一般贸易"
                        elif "征免性质" in cell.value:
                            cell.value = f"征免性质\n一般征税"
                        elif "贸易国" in cell.value:
                            cell.value = f"贸易国(地区)\n印度"
                        elif "运抵国" in cell.value:
                            cell.value = f"运抵国（地区)\n印度"
                        elif "运费" in cell.value:
                            cell.value = f"运费（CNY)\n{fill_dict['运费（CNY)']}"
                        elif "保费" in cell.value:
                            cell.value = f"保费（CNY)\n{fill_dict['保费（CNY)']}"
                        elif "境内发货人" in cell.value:
                            cell.value = f"境内发货人\n{fill_dict['境内发货人']}"   
                        elif "生产销售单位" in cell.value:
                            cell.value = f"生产销售单位\n{fill_dict['生产销售单位']}   "
                        elif "境外收货人" in cell.value:
                            cell.value = f"境外收货人\n{fill_dict['境外收货人']}"
                        elif "合同协议号" in cell.value:
                            cell.value = f"合同协议号\n{fill_dict['合同协议号']}"
        
            wb.save('1.xlsx')

            print("Updated weight and quantity information in 1.xlsx")
    except Exception as e:
        print(f"Error updating 1.xlsx: {e}")

    

    # 处理3.xlsx文件
    if os.path.exists('3.xlsx'):
        from openpyxl import load_workbook
        wb = load_workbook('3.xlsx')
        ws = wb.active
        
        # 遍历前两行查找目标单元格
        for row in range(1, 3):
            for col in range(1, ws.max_column):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if '总货值' in cell.value:
                        ws.cell(row=row, column=col+1, value=t_amount)
                    elif '总净重' in cell.value:
                        ws.cell(row=row, column=col+1, value=t_weight)
        
        wb.save('3.xlsx')
        print("Updated total amount and weight in 3.xlsx")
    
    # 调用merge.py合并文件
    try:
        print("Merging files with merge.py...")
        # 假设1.xlsx和2.xlsx在当前目录下
        # 调用格式：python merge.py 1.xlsx output.xlsx 2.xlsx
        merge_cmd = [sys.executable, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'merge.py'), 
                    '1.xlsx', output_file, '3.xlsx']
        subprocess.run(merge_cmd, check=True)
        print("Files merged successfully!")
        
        # 在Windows系统下自动打开合并后的Excel文件
        if os.name == 'nt':
            merged_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'merged.xlsx')
            if os.path.exists(merged_file):
                os.startfile(merged_file)
                print("Opening merged Excel file...")
    except Exception as e:
        print(f"Error merging files: {e}")
    

    
    # Return the DataFrame for potential further processing or analysis
    return df_output

def main():
    """
    Main function to parse command-line arguments and execute the Excel conversion.
    
    This function sets up an argument parser to handle input, reference, and output
    file paths provided as command-line arguments, then calls the convert_excel function.
    
    Command-line usage:
    python excel_converter.py input.xlsx reference.xlsx output.xlsx
    """
    parser = argparse.ArgumentParser(description='Convert Excel files according to specified format')
    parser.add_argument('input', help='Path to the input Excel file')
    parser.add_argument('reference', help='Path to the reference Excel file')
    parser.add_argument('output', help='Path to save the output Excel file')
    
    args = parser.parse_args()
    
    result = convert_excel(args.input, args.reference, args.output)
    if result is None:
        sys.exit(1)  # Exit with error code if conversion failed

# Entry point of the script
# This conditional ensures the main() function is only executed when the script is run directly,
# not when it's imported as a module (like in the Streamlit app)
if __name__ == "__main__":
    main()